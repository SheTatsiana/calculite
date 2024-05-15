from django.shortcuts import render, redirect, get_object_or_404
from .models import MyCurrentObject, MyObject, Note, Product, WorkDetail
from .forms import MyCurrentObjectForm
from django.urls import reverse
from .forms import ProductForm
from .forms import MyObjectForm
from .forms import NoteForm, WorkDetailForm

from django.http import JsonResponse
from .models import WorkDetail
from django.views.decorators.csrf import csrf_exempt


#url = reverse('users:profile_edit')

def home(request):
    my_current_objects = MyCurrentObject.objects.all()
    return render(request, 'home.html', {'my_current_objects': my_current_objects})

def mycurrentobject(request):
    my_current_objects = MyCurrentObject.objects.all()
    products = Product.objects.all()  # Получаем все продукты
    return render(request, 'mycurrentobject.html', {'my_current_objects': my_current_objects, 'products': products})

#def my_view(request):
    #current_objects = MyCurrentObject.get_objects()
    #return render(request, 'mycurrentobject.html', {'current_objects': current_objects})

def my_object_view(request):
    objects = MyObject.objects.all()  # Получаем все объекты
    return render(request, 'my_object.html', {'objects': objects})

def my_object(request):
    my_objects = MyObject.objects.all()
    return render(request, 'my_object.html', {'my_objects': my_objects})

def note(request):
    notes = Note.objects.all()
    return render(request, 'note.html', {'notes': notes})

def product(request):
    products = Product.objects.all()
    return render(request, 'product.html', {'products': products})

def workdetail(request):
    work_details = WorkDetail.objects.all()
    return render(request, 'workdetail.html', {'work_details': work_details})


def add_mco(request): 
    if request.method == 'POST':
        form = MyCurrentObjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('mycurrentobject')
    else:
        form = MyCurrentObjectForm()
    return render(request, 'add_mco.html', {'form': form})


def add_mo(request):
    if request.method == 'POST':
        form = MyObjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('my_object')
    else:
        form = MyObjectForm()
    return render(request, 'add_mo.html', {'form': form})

def add_note(request):
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('note')
    else:
        form = NoteForm()
    return render(request, 'add_note.html', {'form': form})

def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('product')  # Перенаправляем на страницу со списком продуктов
    else:
        form = ProductForm()
    return render(request, 'add_product.html', {'form': form})

@csrf_exempt
def add_wd(request):
    if request.method == 'POST':
        object_id = request.POST.get('object_id')
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')

        # Создаем новую рабочую деталь
        work_detail = WorkDetail.objects.create(
            my_current_object_id=object_id,
            product_id=product_id,
            quantity=quantity,
        )

        # Возвращаем успешный JSON-ответ
        return JsonResponse({'success': True})

    # Если запрос не POST, возвращаем ошибку
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=400)

def edit_mco(request, pk):
    my_current_object = get_object_or_404(MyCurrentObject, pk=pk)
    if request.method == 'POST':
        form = MyCurrentObjectForm(request.POST, instance=my_current_object)
        if form.is_valid():
            form.save()
            return redirect('mycurrentobject')
    else:
        form = MyCurrentObjectForm(instance=my_current_object)
    return render(request, 'edit_mco.html', {'form': form})

def edit_mo(request, pk):
    my_object = get_object_or_404(MyObject, pk=pk)
    if request.method == 'POST':
        form = MyObjectForm(request.POST, instance=my_object)
        if form.is_valid():
            form.save()
            return redirect('my_object')
    else:
        form = MyObjectForm(instance=my_object)
    
    # Передаем название объекта в контекст шаблона
    object_name = my_object.name
    context = {'form': form, 'object_name': object_name}
    return render(request, 'edit_mo.html', context)

def edit_note(request, pk):
    note = get_object_or_404(Note, pk=pk)
    if request.method == 'POST':
        form = NoteForm(request.POST, instance=note)
        if form.is_valid():
            form.save()
            return redirect('note')
    else:
        form = NoteForm(instance=note)
    return render(request, 'edit_note.html', {'form': form})

def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product')  # Перенаправляем на страницу со списком продуктов
    else:
        form = ProductForm(instance=product)
    return render(request, 'edit_product.html', {'form': form, 'product': product})

def edit_wd(request, pk):
    work_detail = get_object_or_404(WorkDetail, pk=pk)
    if request.method == 'POST':
        form = WorkDetailForm(request.POST, instance=work_detail)
        if form.is_valid():
            form.save()
            return redirect('workdetail')  
    else:
        form = WorkDetailForm(instance=work_detail)
    return render(request, 'edit_wd.html', {'form': form, 'work_detail': work_detail})



def delete_product(request, pk):
    # Получаем объект продукта или возвращаем ошибку 404, если продукт не найден
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        # При получении POST-запроса удаляем продукт
        product.delete()
        # Перенаправляем пользователя на страницу со списком продуктов
        return redirect('product')
    # Возвращаем шаблон для подтверждения удаления продукта
    return render(request, 'delete_product.html', {'product': product})

def delete_mo(request, pk):
    my_object = get_object_or_404(MyObject, pk=pk)
    if request.method == 'POST':
        my_object.delete()
        return redirect('my_object')
    return render(request, 'delete_mo.html', {'my_object': my_object})


def delete_note(request, pk):
    note = get_object_or_404(Note, pk=pk)
    if request.method == 'POST':
        note.delete()
        return redirect('note')
    return render(request, 'delete_note.html', {'note': note})


def delete_wd(request, work_detail_id):
    work_detail = get_object_or_404(WorkDetail, pk=work_detail_id)
    if request.method == 'POST':
        work_detail.delete()
        return redirect('workdetail')  # Redirect to the list view
    return render(request, 'delete_wd.html', {'work_detail': work_detail})



import openpyxl
from django.http import HttpResponse

def export_to_excel(request):
    products = Product.objects.all()

    # Создаем новый документ Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    # Заполняем заголовки столбцов
    ws.append(["Наименование", "Цена", "Описание", "Фото"])

    # Заполняем данные о продуктах
    for product in products:
        ws.append([product.name, product.price, product.description, product.image.url if product.image else ""])

    # Создаем HTTP-ответ с содержимым файла Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=catalog.xlsx'

    # Сохраняем документ Excel в HTTP-ответ
    wb.save(response)

    return response



def export_current_objects_to_excel(request):
    my_current_objects = MyCurrentObject.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Текущие проекты"

    ws.append(["Название", "Адрес", "Начало работ", "Окончание работ", "Заказчик", "Исполнитель"])

    for my_current_object in my_current_objects:
        ws.append([
            my_current_object.my_object.name,
            my_current_object.my_object.address,
            my_current_object.my_object.start_date,
            my_current_object.my_object.end_date,
            my_current_object.my_object.customer_name,
            my_current_object.my_object.executor_name
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=current_objects.xlsx'

    wb.save(response)

    return response



def export_all_objects_to_excel(request):
    my_objects = MyObject.objects.all()

    # Создаем новый документ Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все проекты"

    # Заполняем заголовки столбцов
    ws.append(["Название", "Адрес", "Начало работ", "Окончание работ", "Заказчик", "Исполнитель"])

    # Заполняем данные о объектах
    for my_object in my_objects:
        ws.append([
            my_object.name,
            my_object.address,
            my_object.start_date,
            my_object.end_date,
            my_object.customer_name,
            my_object.executor_name
        ])

    # Создаем HTTP-ответ с содержимым файла Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_objects.xlsx'

    # Сохраняем документ Excel в HTTP-ответ
    wb.save(response)

    return response


def mycurrentobject(request):
    my_current_objects = MyCurrentObject.objects.all()
    return render(request, 'mycurrentobject.html', {'my_objects': my_current_objects})


def delete_mco(request, pk):
    my_current_object = get_object_or_404(MyCurrentObject, pk=pk)
    if request.method == 'POST':
        my_current_object.delete()  
        return redirect('mycurrentobject')
    return render(request, 'delete_mco.html', {'my_current_object': my_current_object})